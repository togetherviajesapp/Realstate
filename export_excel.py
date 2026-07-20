#!/usr/bin/env python3
"""
Genera/actualiza el Excel de reporte semanal a partir de data.json (fuente unica
de verdad, la misma que usa el sitio web). Se corre despues de process.py en
cada corrida automatica.

Uso:
    python export_excel.py --data data.json --excel Reporte_Inmuebles_Montevideo.xlsx
"""
import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import json

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BASE_FONT = Font(name=FONT_NAME)
BOLD_FONT = Font(name=FONT_NAME, bold=True)

LISTADO_COLS = [
    "portal", "operacion", "tipo_inmueble", "barrio", "dormitorios", "banos", "m2",
    "precio_moneda", "precio_valor", "gastos_comunes", "titulo", "url",
    "primera_vez_visto", "ultima_vez_visto", "estado",
]
LISTADO_WIDTHS = [12, 10, 14, 16, 11, 8, 8, 10, 12, 12, 40, 45, 14, 14, 12]

BAJAS_COLS = [
    "portal", "operacion", "tipo_inmueble", "barrio", "precio_moneda", "precio_valor",
    "titulo", "url", "primera_vez_visto", "ultima_vez_visto", "fecha_baja", "dias_publicado",
]
BAJAS_WIDTHS = [12, 10, 14, 16, 10, 12, 40, 45, 14, 14, 12, 12]

HIST_COLS = ["fecha", "portal", "operacion", "nuevos", "mantenidos", "bajas", "total_activos"]


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(wb, name, cols, rows, widths):
    ws = wb.create_sheet(name)
    ws.append(cols)
    for r in rows:
        vals = []
        for c in cols:
            v = r.get(c, "")
            if c in ("precio_valor", "gastos_comunes", "dormitorios", "banos", "m2", "dias_publicado") and v not in ("", None):
                try:
                    v = float(v)
                    if v == int(v):
                        v = int(v)
                except (ValueError, TypeError):
                    pass
            vals.append(v)
        ws.append(vals)
        for c in range(1, len(cols) + 1):
            ws.cell(row=ws.max_row, column=c).font = BASE_FONT
    style_header(ws, len(cols))
    autosize(ws, widths)
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True, help="ruta de data.json (se lee, no se modifica)")
    ap.add_argument("--excel", required=True, help="ruta del Excel a generar/sobreescribir")
    args = ap.parse_args()

    with open(args.data, encoding="utf-8") as f:
        data = json.load(f)

    listado = data.get("listado", [])
    bajas = data.get("bajas_historico", [])
    historico = data.get("historico", [])
    fecha = data.get("actualizado", "")
    resumen = data.get("resumen_ultima_corrida", {})

    wb = Workbook()
    wb.remove(wb.active)

    # --- Instrucciones ---
    ws = wb.create_sheet("Instrucciones", 0)
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Reporte semanal de inmuebles - Montevideo", True),
        ("", False),
        ("Este archivo se genera automaticamente cada semana a partir de la misma", False),
        ("informacion que muestra el sitio web (InfoCasas, Gallito y MercadoLibre).", False),
        ("", False),
        ("Hojas:", True),
        ("- Listado actual: todos los avisos activos hoy, con fecha de primera y ultima vez visto.", False),
        ("- Posibles bajas: avisos que dejaron de verse en el portal (posible venta/alquiler concretado,", False),
        ("  o baja del aviso por otro motivo -- no se puede confirmar cual de las dos cosas paso).", False),
        ("- Historico semanal: log de conteos (nuevos/mantenidos/bajas/activos) por portal y operacion.", False),
        ("- Resumen y Grafica: totales por semana (formulas SUMIFS) y graficos de evolucion.", False),
        ("", False),
        (f"Ultima actualizacion: {fecha}", False),
        (f"Activos: {data.get('total_activos', '')}  |  Nuevos esta corrida: {resumen.get('nuevos', '')}  |  "
         f"Mantenidos: {resumen.get('mantenidos', '')}  |  Bajas esta corrida: {resumen.get('bajas', '')}", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = BOLD_FONT if bold else BASE_FONT

    # --- Listado actual ---
    write_table(wb, "Listado actual", LISTADO_COLS, listado, LISTADO_WIDTHS)

    # --- Posibles bajas ---
    write_table(wb, "Posibles bajas", BAJAS_COLS, bajas, BAJAS_WIDTHS)

    # --- Historico semanal ---
    ws = wb.create_sheet("Historico semanal")
    ws.append(HIST_COLS)
    for r in historico:
        ws.append([r.get(c, "") for c in HIST_COLS])
        for c in range(1, len(HIST_COLS) + 1):
            ws.cell(row=ws.max_row, column=c).font = BASE_FONT
    style_header(ws, len(HIST_COLS))
    autosize(ws, [12, 12, 12, 10, 12, 8, 14])

    # --- Resumen y Grafica ---
    ws = wb.create_sheet("Resumen y Grafica")
    fechas = sorted({r["fecha"] for r in historico})
    resumen_cols = ["fecha", "nuevos_venta", "nuevos_alquiler", "activos_venta", "activos_alquiler", "bajas_venta", "bajas_alquiler"]
    ws.append(resumen_cols)
    for i, f in enumerate(fechas):
        r = i + 2
        ws.cell(row=r, column=1, value=f)
        ws.cell(row=r, column=2, value=f"=SUMIFS('Historico semanal'!D:D,'Historico semanal'!A:A,A{r},'Historico semanal'!C:C,\"venta\")")
        ws.cell(row=r, column=3, value=f"=SUMIFS('Historico semanal'!D:D,'Historico semanal'!A:A,A{r},'Historico semanal'!C:C,\"alquiler\")")
        ws.cell(row=r, column=4, value=f"=SUMIFS('Historico semanal'!G:G,'Historico semanal'!A:A,A{r},'Historico semanal'!C:C,\"venta\")")
        ws.cell(row=r, column=5, value=f"=SUMIFS('Historico semanal'!G:G,'Historico semanal'!A:A,A{r},'Historico semanal'!C:C,\"alquiler\")")
        ws.cell(row=r, column=6, value=f"=SUMIFS('Historico semanal'!F:F,'Historico semanal'!A:A,A{r},'Historico semanal'!C:C,\"venta\")")
        ws.cell(row=r, column=7, value=f"=SUMIFS('Historico semanal'!F:F,'Historico semanal'!A:A,A{r},'Historico semanal'!C:C,\"alquiler\")")
        for c in range(1, 8):
            ws.cell(row=r, column=c).font = BASE_FONT
    style_header(ws, len(resumen_cols))
    autosize(ws, [12, 14, 14, 14, 14, 12, 12])
    ws.cell(row=1, column=9, value="Nota: columnas B-G se calculan con SUMIFS sobre 'Historico semanal'. No editar a mano.").font = Font(name=FONT_NAME, italic=True, size=9)

    last_data_row = 1 + len(fechas)
    if last_data_row >= 2:
        chart = LineChart()
        chart.title = "Evolucion de inmuebles activos (Montevideo)"
        chart.style = 2
        chart.y_axis.title = "Cantidad de avisos activos"
        chart.x_axis.title = "Corrida"
        data_ref = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=last_data_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 24
        chart.height = 12
        ws.add_chart(chart, "I3")

        chart2 = LineChart()
        chart2.title = "Nuevos ingresos por corrida (Montevideo)"
        chart2.style = 10
        chart2.y_axis.title = "Nuevos avisos"
        chart2.x_axis.title = "Corrida"
        data2 = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last_data_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.width = 24
        chart2.height = 12
        ws.add_chart(chart2, "I28")

    out_dir = os.path.dirname(args.excel)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(args.excel)
    print(f"OK: {args.excel} generado. Activos={len(listado)} Bajas historicas={len(bajas)} "
          f"Filas historico={len(historico)}")


if __name__ == "__main__":
    main()
